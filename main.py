from openpyxl import Workbook, load_workbook

class Customer:
    def __init__(self, name, email, phone):
        self.name = name
        self.email = email
        self.phone = phone

    def __str__(self):
        return f"Name: {self.name}\nEmail: {self.email}\nPhone: {self.phone}"

    def to_dict(self):
        return {
            "Name": self.name,
            "Email": self.email,
            "Phone": self.phone
        }

class CustomerManager:
    def __init__(self, file_path):
        self.file_path = file_path
        self.customers = []

    def add_customer(self, customer):
        self.customers.append(customer)

    def save_to_excel(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Customers"

        headers = ["Name", "Email", "Phone"]
        sheet.append(headers)

        for customer in self.customers:
            customer_data = customer.to_dict()
            sheet.append(customer_data.values())

        workbook.save(self.file_path)
        print("Data saved to Excel file.")

    def load_from_excel(self):
        try:
            workbook = load_workbook(self.file_path)
            sheet = workbook.active

            self.customers = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, email, phone = row
                customer = Customer(name, email, phone)
                self.customers.append(customer)

            print("Data loaded from Excel file.")
        except FileNotFoundError:
            print("Excel file not found.")
        except Exception as e:
            print("An error occurred while loading data:", str(e))


def get_user_input():
    name = input("Enter customer name: ")
    email = input("Enter customer email: ")
    phone = input("Enter customer phone number: ")
    return name, email, phone


# Usage Example:

file_path = "customers.xlsx"

manager = CustomerManager(file_path)

while True:
    print("1. Add Customer")
    print("2. Save to Excel")
    print("3. Load from Excel")
    print("4. Display Customers")
    print("5. Exit")

    choice = input("Enter your choice: ")

    if choice == "1":
        name, email, phone = get_user_input()
        customer = Customer(name, email, phone)
        manager.add_customer(customer)
    elif choice == "2":
        manager.save_to_excel()
    elif choice == "3":
        manager.load_from_excel()
    elif choice == "4":
        for customer in manager.customers:
            print(customer)
            print()
    elif choice == "5":
        break
    else:
        print("Invalid choice. Please try again.")

